#!/usr/bin/env python3
"""
PortfolioIQ v1.0
Grant Financial Summary Action Engine

Purpose:
PortfolioIQ reviews grant Financial Summary reports, applies sponsor-aware
financial logic, flags risk, and generates standardized action guidance
and manager comments.

Save this file as:
C:\GrantActions\PortfolioIQ.py
"""

from __future__ import annotations
import os, re, sys
from datetime import datetime
from typing import Dict, Tuple, Optional, List
from openpyxl import load_workbook

# =========================
# SETTINGS
# =========================
DEFAULT_FA_RATE = 0.60
SUBAWARD_FIRST_CAP = 25000.0

# =========================
# STARTUP BANNER
# =========================
def startup_banner():
    print("=" * 60)
    print("PortfolioIQ v1.0")
    print("Grant Financial Summary Action Engine")
    print("=" * 60)

# =========================
# HELPERS
# =========================
def to_number(x) -> float:
    try:
        return float(str(x).replace(",", "").replace("-", "")) if x else 0.0
    except:
        return 0.0

def parse_fa_rate(text: str) -> Optional[float]:
    if not text:
        return None
    m = re.search(r"(\d{1,3})\s*%", str(text))
    if m:
        return float(m.group(1)) / 100
    return None

def find_fa_rate(ws, row: int) -> Optional[float]:
    for r in range(row - 1, max(1, row - 120), -1):
        for c in range(1, 8):
            v = ws.cell(r, c).value
            if v and "F&A" in str(v):
                return parse_fa_rate(v)
    return None

# =========================
# CORE ENGINE
# =========================
def run_portfolioiq(ws):
    header_row = None
    for r in range(1, 50):
        if ws.cell(r, 1).value == "Grant":
            header_row = r
            break
    if not header_row:
        raise RuntimeError("Header row not found")

    last_row = ws.max_row
    updated = 0

    for r in range(header_row + 1, last_row + 1):
        sclass = str(ws.cell(r, 3).value).upper()
        if "CAPITAL" in sclass:
            budget = to_number(ws.cell(r, 4).value)
            actual = to_number(ws.cell(r, 5).value)
            rate = find_fa_rate(ws, r) or DEFAULT_FA_RATE
            impact = (budget - actual) * rate
            ws.cell(r, 10).value = impact
            updated += 1

    return updated

# =========================
# MAIN
# =========================
def main():
    startup_banner()

    if len(sys.argv) < 2:
        print("ERROR: No Excel file provided.")
        return

    input_file = sys.argv[1]
    print(f"Loading workbook: {input_file}")

    wb = load_workbook(input_file)
    ws = wb.active

    print("Applying PortfolioIQ rules...")
    updated = run_portfolioiq(ws)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = input_file.replace(".xlsx", f"_PortfolioIQ_{timestamp}.xlsx")
    wb.save(output_file)

    print("✔ PortfolioIQ completed successfully.")
    print(f"   Rows updated: {updated}")
    print(f"   Saved file: {output_file}")

if __name__ == "__main__":
    main()
